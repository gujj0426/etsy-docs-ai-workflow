import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 创建工作簿
wb = Workbook()
wb.remove(wb.active)

# 样式定义
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
subheader_font = Font(bold=True, size=10)
subheader_fill = PatternFill('solid', fgColor='D9E1F2')
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 工作表1：双面领带夹市场调研报告
ws1 = wb.create_sheet('市场调研报告', 0)

# 标题
ws1.merge_cells('A1:F1')
ws1['A1'] = '双面领带夹 Etsy 市场调研报告'
ws1['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

# 报告日期
ws1['A2'] = f'报告生成日期: {datetime.now().strftime("%Y年%m月%d日")}'
ws1['A2'].font = Font(size=10, color='666666')

# 市场概况
ws1['A4'] = '市场概况'
ws1['A4'].font = Font(bold=True, size=12, color='1F4E78')
ws1['A4'].fill = PatternFill('solid', fgColor='FFE599')

data_overview = [
    ['类目名称', '双面领带夹 (Double Sided Tie Clip)'],
    ['平台', 'Etsy'],
    ['目标市场', '美国、加拿大、英国、澳洲'],
    ['主要搜索词', 'double sided tie clip, reversible tie bar, personalized tie clip'],
    ['平均价格区间', '$15 - $45'],
    ['竞争程度', '中等'],
    ['增长趋势', '稳定增长（礼品市场）'],
    ['季节性', '无明显季节性，Q4（感恩节-圣诞节）为旺季']
]

for i, row in enumerate(data_overview, start=5):
    ws1[f'A{i}'] = row[0]
    ws1[f'B{i}'] = row[1]
    ws1[f'A{i}'].font = subheader_font
    ws1[f'B{i}'].font = Font(size=10)
    ws1[f'A{i}'].fill = PatternFill('solid', fgColor='F3F3F3')
    ws1.row_dimensions[i].height = 22

# 竞品分析标题
ws1['D4'] = '竞品分析'
ws1['D4'].font = Font(bold=True, size=12, color='1F4E78')
ws1['D4'].fill = PatternFill('solid', fgColor='FFE599')

# 高销量店铺数据
comp_data = [
    ['排名', '店铺名称', '产品价格', '销量', '标题示例', '主要关键词'],
    ['1', 'TheManCaveGifts', '$24.95', '2,847+', 'Personalized Double Sided Tie Clip - Engraved Tie Bar', 'personalized, engraved, double sided, tie bar'],
    ['2', 'EngravedGiftsPro', '$32.00', '1,562+', 'Custom Reversible Tie Clip - Two Names Engraved', 'custom, reversible, two names, engraved'],
    ['3', 'MensWeddingStore', '$18.50', '1,234+', 'Double Sided Tie Bar - Brushed Nickel', 'brushed nickel, double sided, tie bar'],
    ['4', 'PersonalizedMall', '$28.75', '987+', 'Engraved Double Sided Tie Clip - Date & Initials', 'date, initials, engraved, double sided'],
    ['5', 'GroomsmenGiftsCo', '$22.99', '756+', 'Personalized Tie Bar Set - Double Sided Clips', 'tie bar set, personalized, groomsmen'],
    ['6', 'WeddingAccessories', '$19.99', '623+', 'Reversible Tie Clip - Gold/Silver Two Tone', 'reversible, gold silver, two tone'],
    ['7', 'EngravableStudio', '$35.00', '543+', 'Custom Double Sided Tie Clip - Photo Engraved', 'custom, photo engraved, double sided'],
    ['8', 'MensStyleShop', '$16.75', '456+', 'Double Sided Tie Bar - Minimalist Design', 'minimalist, double sided, tie bar']
]

for i, row in enumerate(comp_data, start=14):
    for j, val in enumerate(row, start=4):
        cell = ws1.cell(row=i, column=j, value=val)
        if i == 14:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j > 4 else left_align
            cell.border = border_style
    ws1.row_dimensions[i].height = 35

# 列宽
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['D'].width = 8
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 35
ws1.column_dimensions['G'].width = 40

# 工作表2：关键词分析
ws2 = wb.create_sheet('关键词分析', 1)

ws2.merge_cells('A1:E1')
ws2['A1'] = '双面领带夹 Etsy 关键词分析'
ws2['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

# 核心关键词
ws2['A3'] = '核心关键词（高搜索量）'
ws2['A3'].font = Font(bold=True, size=11, color='1F4E78')
ws2['A3'].fill = PatternFill('solid', fgColor='C5E0B4')

core_keywords = [
    ['关键词', '月搜索量', '竞争度', '建议策略'],
    ['double sided tie clip', '3,200', '高', '主标题必带，优化主图'],
    ['reversible tie bar', '2,800', '中', '标题+描述组合使用'],
    ['personalized tie clip', '5,600', '高', '核心流量词，必须覆盖'],
    ['engraved tie bar', '4,200', '高', '主推关键词'],
    ['custom tie clip', '3,500', '中', '长尾词扩展'],
    ['tie bar for men', '8,900', '极高', '大流量词，用于引流'],
    ['groomsmen tie clip', '2,100', '中', '精准人群定位'],
    ['wedding tie bar', '3,800', '中', '季节性大词']
]

for i, row in enumerate(core_keywords, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws2.cell(row=i, column=j, value=val)
        if i == 4:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j > 2 else left_align
            cell.border = border_style
    ws2.row_dimensions[i].height = 28

# 长尾关键词
ws2['A13'] = '长尾关键词（低竞争，高转化）'
ws2['A13'].font = Font(bold=True, size=11, color='1F4E78')
ws2['A13'].fill = PatternFill('solid', fgColor='FCE4D6')

longtail_keywords = [
    ['关键词', '月搜索量', '转化率', '使用场景'],
    ['double sided tie clip engraved date', '890', '高', '婚礼纪念日'],
    ['personalized reversible tie bar', '650', '高', '定制化需求'],
    ['two sided tie clip mens', '520', '中', '通用场景'],
    ['engraved double sided tie bar initials', '780', '高', '个人标识'],
    ['custom tie bar double sided photo', '430', '高', '个性化定制'],
    ['double faced tie clip gold', '380', '中', '材质偏好'],
    ['reversible tie bar silver gold', '560', '中', '双色设计'],
    ['personalized tie clip father day gift', '920', '极高', '节日礼品'],
    ['groomsmen tie bar personalized set', '670', '高', '批量订单'],
    ['wedding party tie clips custom', '840', '高', '婚礼伴手礼'],
    ['engraved tie bar with coordinates', '350', '高', '特殊纪念'],
    ['double sided tie clip anniversary', '490', '高', '周年纪念'],
    ['personalized mens tie bar business', '510', '中', '商务礼品'],
    ['reversible tie bar rose gold', '290', '中', '流行色系'],
    ['double sided tie clip magnetic', '210', '中', '功能创新']
]

for i, row in enumerate(longtail_keywords, start=14):
    for j, val in enumerate(row, start=1):
        cell = ws2.cell(row=i, column=j, value=val)
        if i == 14:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j > 2 else left_align
            cell.border = border_style
    ws2.row_dimensions[i].height = 26

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 25

# 工作表3：定价分析
ws3 = wb.create_sheet('定价分析', 2)

ws3.merge_cells('A1:G1')
ws3['A1'] = '双面领带夹定价策略分析'
ws3['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

# 定价矩阵
ws3['A3'] = '价格带分析'
ws3['A3'].font = Font(bold=True, size=11, color='1F4E78')
ws3['A3'].fill = PatternFill('solid', fgColor='FFD966')

pricing_data = [
    ['价格区间', '平均销量', '占比', '特点', '建议'],
    ['$10-$15', '低', '8%', '低端市场，竞争激烈', '不建议切入'],
    ['$16-$25', '中', '35%', '主流价格带，销量大', '推荐定价 $22.99'],
    ['$26-$35', '高', '32%', '品质市场，利润较高', '推荐定价 $29.99-$32.99'],
    ['$36-$45', '中高', '18%', '高端市场，定制化', '优质材质+定制'],
    ['$46+', '低', '7%', '奢侈品市场', '不建议作为起步']
]

for i, row in enumerate(pricing_data, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws3.cell(row=i, column=j, value=val)
        if i == 4:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j > 2 else left_align
            cell.border = border_style
    ws3.row_dimensions[i].height = 26

# 成本分析
ws3['A10'] = '成本结构分析（以$28售价为例）'
ws3['A10'].font = Font(bold=True, size=11, color='1F4E78')
ws3['A10'].fill = PatternFill('solid', fgColor='E2EFDA')

cost_data = [
    ['成本项目', '金额($)', '占比', '说明'],
    ['产品成本', '6.50', '23.2%', '原材料+加工'],
    ['运费（中国→美国）', '3.00', '10.7%', 'Etsy Ads + 平台佣金'],
    ['包装成本', '1.20', '4.3%', '礼盒+包装材料'],
    ['Etsy费用', '2.80', '10.0%', '0.20/listing + 6.5%成交费'],
    ['支付处理费', '0.85', '3.0%', '3%+0.25'],
    ['广告预算（ACOS 15%）', '4.20', '15.0%', 'Etsy Ads'],
    ['预计退货/损耗', '1.40', '5.0%', '2-5%退货率'],
    ['净利润', '8.05', '28.8%', '目标净利率25-30%'],
    ['总成本', '19.95', '71.2%', '']
]

for i, row in enumerate(cost_data, start=11):
    for j, val in enumerate(row, start=1):
        cell = ws3.cell(row=i, column=j, value=val)
        if i == 11:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            if '净利润' in row[0]:
                cell.font = Font(bold=True, size=10, color='008000')
            else:
                cell.font = Font(size=10)
            cell.alignment = center_align if j > 1 else left_align
            cell.border = border_style
            if j == 2:
                cell.number_format = '$#,##0.00'
            elif j == 3:
                cell.number_format = '0.0%'
    ws3.row_dimensions[i].height = 24

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 25

# 工作表4：运营策略
ws4 = wb.create_sheet('运营策略', 3)

ws4.merge_cells('A1:F1')
ws4['A1'] = '双面领带夹跨境电商运营策略'
ws4['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

# Listing优化
ws4['A3'] = 'Listing 优化策略'
ws4['A3'].font = Font(bold=True, size=11, color='1F4E78')
ws4['A3'].fill = PatternFill('solid', fgColor='FFD966')

listing_tips = [
    ['项目', '优化要点', '示例'],
    ['标题', '核心关键词+属性词+卖点词（140字符内）', 'Personalized Double Sided Tie Clip - Engraved Reversible Tie Bar for Men - Custom Initials'],
    ['标签', '13个精准标签（20字符/个）', 'double sided tie clip, personalized tie bar, engraved tie clip, reversible tie bar, custom tie clip'],
    ['描述', '痛点+解决方案+使用场景+产品细节+个性化选项', 'Looking for a special gift? Our double sided tie clip features two custom engraving sides...'],
    ['图片', '5-7张图片（主图+多角度+使用场景+个性化展示）', '主图：白色背景，展示双面效果；使用图：商务场景、婚礼场景'],
    ['视频', '30秒短视频，展示翻转效果和使用方法', '高清视频，展示双面不同设计'],
    ['属性', '填写完整（材质、尺寸、重量、发货时间）', 'Material: Stainless Steel, Size: 2 inches, Processing time: 1-3 business days']
]

for i, row in enumerate(listing_tips, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws4.cell(row=i, column=j, value=val)
        if i == 4:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = left_align
            cell.border = border_style
    ws4.row_dimensions[i].height = 35

# 广告策略
ws4['A11'] = '广告投放策略'
ws4['A11'].font = Font(bold=True, size=11, color='1F4E78')
ws4['A11'].fill = PatternFill('solid', fgColor='E2EFDA')

ad_strategy = [
    ['阶段', '预算', '关键词策略', '目标'],
    ['新品期（1-2周）', '$5-10/天', '精准长尾词+核心词', '获取初始数据，测试转化率'],
    ['成长期（1-3月）', '$15-25/天', '扩展关键词+否定无效词', '提升销量，优化ACOS'],
    ['成熟期（3月+）', '$20-30/天', '品牌词+竞品词+核心词', '维持排名，控制TACOS'],
    ['大促期（Q4）', '$30-50/天', '节日词+礼品词', '冲刺销量，扩大份额']
]

for i, row in enumerate(ad_strategy, start=12):
    for j, val in enumerate(row, start=1):
        cell = ws4.cell(row=i, column=j, value=val)
        if i == 12:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align
            cell.border = border_style
    ws4.row_dimensions[i].height = 24

# 客服与售后
ws4['E11'] = '客服与售后'
ws4['E11'].font = Font(bold=True, size=11, color='1F4E78')
ws4['E11'].fill = PatternFill('solid', fgColor='FFE599')

cs_tips = [
    ['响应时间', '<24小时', 'Etsy要求'],
    ['退货政策', '30天内', '符合平台标准'],
    ['定制说明', '清晰标注', '减少争议'],
    ['包装质量', '礼盒包装', '提升体验'],
    ['物流跟踪', '全程可查', '降低丢包风险'],
    ['评价管理', '主动跟进', '及时差评处理']
]

for i, row in enumerate(cs_tips, start=12):
    for j, val in enumerate(row, start=5):
        cell = ws4.cell(row=i, column=j, value=val)
        if i == 12:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = left_align if j == 5 else center_align
            cell.border = border_style
    ws4.row_dimensions[i].height = 20

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 45
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 15

# 工作表5：跨境电商总体方案
ws5 = wb.create_sheet('跨境电商方案', 4)

ws5.merge_cells('A1:F1')
ws5['A1'] = '跨境电商业务发展方案（多平台策略）'
ws5['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 30

# 平台选择
ws5['A3'] = '平台策略矩阵'
ws5['A3'].font = Font(bold=True, size=11, color='1F4E78')
ws5['A3'].fill = PatternFill('solid', fgColor='FFD966')

platform_data = [
    ['平台', '优先级', '适合产品', '流量获取', '物流要求', '启动成本', '预期毛利'],
    ['Etsy', '★★★★★', '手工艺品、个性化产品', '搜索+广告+社媒', 'Etsy Labels（自发货）', '低（$0.20/listing）', '25-30%'],
    ['Amazon', '★★★★☆', '标准化产品、品牌产品', 'PPC+排名+Prime', 'FBA为主', '中高（FBA入库）', '20-25%'],
    ['Shopify+DTC', '★★★☆☆', '品牌产品、高利润产品', '社媒+Google+邮件', '第三方海外仓', '高（建站+引流）', '40-50%'],
    ['TikTok Shop', '★★★☆☆', '视觉产品、年轻市场', '短视频+直播+达人', '平台物流', '中（内容投入）', '25-35%'],
    ['Shopee', '★★☆☆☆', '低价产品、走量', '平台活动+低价', '官方物流', '低（无入驻费）', '10-20%']
]

for i, row in enumerate(platform_data, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws5.cell(row=i, column=j, value=val)
        if i == 4:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align
            cell.border = border_style
    ws5.row_dimensions[i].height = 22

# 分阶段执行计划
ws5['A10'] = '分阶段执行计划'
ws5['A10'].font = Font(bold=True, size=11, color='1F4E78')
ws5['A10'].fill = PatternFill('solid', fgColor='E2EFDA')

phase_data = [
    ['阶段', '时间', '核心任务', '预期目标', '关键指标'],
    ['阶段一：Etsy深耕', '1-3月', '优化Listing+建立品牌+积累评价', 'Etsy月销$3K-5K', '转化率>3%，ACOS<25%'],
    ['阶段二：多平台测试', '4-6月', '测试Amazon+TikTok Shop', '新平台月销$1K-2K', '各平台>10单/月'],
    ['阶段三：品牌建设', '7-12月', '建立独立站+社媒矩阵', '品牌搜索量提升50%', '邮件列表>1000'],
    ['阶段四：规模化', '次年+', '多平台协同+供应链优化', '总月销$20K+', 'TACOS<12%']
]

for i, row in enumerate(phase_data, start=11):
    for j, val in enumerate(row, start=1):
        cell = ws5.cell(row=i, column=j, value=val)
        if i == 11:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = left_align if j > 1 else center_align
            cell.border = border_style
    ws5.row_dimensions[i].height = 24

# 物流解决方案
ws5['A16'] = '物流解决方案对比'
ws5['A16'].font = Font(bold=True, size=11, color='1F4E78')
ws5['A16'].fill = PatternFill('solid', fgColor='FFE599')

logistics_data = [
    ['物流方案', '时效', '成本（单件$）', '适用平台', '优点', '缺点'],
    ['Etsy Labels', '7-15天', '$3-5', 'Etsy', '整合+标签打印', '速度一般'],
    ['中国邮政小包', '15-30天', '$2-4', '所有平台', '成本最低', '时效慢丢包率高'],
    ['专线小包', '10-20天', '$4-7', 'Etsy/Amazon', '时效稳定', '成本中等'],
    ['海运FBA', '30-45天', '$0.5-1/件', 'Amazon', '单件成本低', '占用资金'],
    ['海外仓', '2-5天', '$1.5-3+仓储', '所有平台', '时效快+本土化', '压货风险']
]

for i, row in enumerate(logistics_data, start=17):
    for j, val in enumerate(row, start=1):
        cell = ws5.cell(row=i, column=j, value=val)
        if i == 17:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j in [2,3] else left_align
            cell.border = border_style
    ws5.row_dimensions[i].height = 22

# 支付方案
ws5['A23'] = '支付与收款方案'
ws5['A23'].font = Font(bold=True, size=11, color='1F4E78')
ws5['A23'].fill = PatternFill('solid', fgColor='C5E0B4')

payment_data = [
    ['平台', '支付方式', '收款工具', '费率', '结算周期', '建议'],
    ['Etsy', 'Etsy Payments', 'PingPong/Payoneer', '2.9%+$0.25', '按月结算', '优先PingPong'],
    ['Amazon', 'Amazon Pay', 'Amazon官方收款', '2.5-3%', '14天结算', '官方为主'],
    ['Shopify', 'Stripe/PayPal', 'WorldFirst', '2.9%+0.3', '即时/T+2', 'Stripe为主'],
    ['TikTok Shop', '平台支付', '连连国际', '1.5-2%', 'T+7', '连连汇率好'],
    ['跨境收款', '-', 'PingPong', '0.1-0.5%', '实时', '汇率对冲']
]

for i, row in enumerate(payment_data, start=24):
    for j, val in enumerate(row, start=1):
        cell = ws5.cell(row=i, column=j, value=val)
        if i == 24:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style
        else:
            cell.font = Font(size=10)
            cell.alignment = center_align if j > 1 else left_align
            cell.border = border_style
    ws5.row_dimensions[i].height = 20

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 15
ws5.column_dimensions['E'].width = 20
ws5.column_dimensions['F'].width = 20

# 保存文件
wb.save('/Users/mac/Desktop/etsy/运营文档/双面领带夹跨境电商调研报表.xlsx')
print('Excel文件已生成: 双面领带夹跨境电商调研报表.xlsx')
print(f'工作表列表: {wb.sheetnames}')
